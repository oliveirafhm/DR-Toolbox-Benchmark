import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

root = tk.Tk()
root.withdraw()
root.update()
file_path = filedialog.askopenfilename()

root.destroy()

wb = load_workbook(file_path)
ws = wb.active
#print(ws['J1'].value)

scenario_col = 'J'
xls_file_name_col = 'C'

for i in range(2,ws.max_row):
    scenario_value = str(ws[xls_file_name_col+str(i)].value)
    new_scenario_value = scenario_value.split('-',1)[1].split('.',1)[0]
    ws[scenario_col+str(i)] = new_scenario_value
    

